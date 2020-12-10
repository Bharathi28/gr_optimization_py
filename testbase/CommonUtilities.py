import os
import pathlib
from datetime import date
from openpyxl import load_workbook
import xlsxwriter
import xlrd
import xlwt
from datetime import datetime


def getExcelData(input_path, sheet_name):
    book = xlrd.open_workbook(input_path)
    sheet = book.sheet_by_name(sheet_name)

    # data = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    # print(data)
    # return data

    for i in range(1, sheet.nrows):
        print(i)


def populateoutputexcel():
    today = date.today()
    Current_date = today.strftime("%m%d%Y")  # d3 = today.strftime("%m/%d/%y")
    print("Current Date : ", Current_date)
    SheetName = 'CrepeErase'
    ExcelName = 'Buyflow'
    script_dir = os.path.dirname(__file__)  # <-- absolute dir the script is in
    rel_path = "Input_Output/Output/" + ExcelName + "_" + Current_date + ".xlsx"
    abs_file_path = os.path.join(script_dir, rel_path)

    print(abs_file_path)
    excel_exist = pathlib.Path(abs_file_path)

    if excel_exist.exists():
        print("Existing file Name : " + abs_file_path)
        wb = load_workbook(abs_file_path, read_only=True)  # open an Excel file and return a workbook

        if SheetName in wb.sheetnames:
            print(SheetName + ' sheet is exists')
        else:
            worksheet = workbook.add_worksheet(SheetName)
    else:
        print("New file Name : " + abs_file_path)
        # Create a workbook and add a worksheet.
        workbook = xlsxwriter.Workbook(abs_file_path)
        worksheet = workbook.add_worksheet(SheetName)

        # Some data we want to write to the worksheet.
        # expenses = (
        #     ['Rent', 1000],
        #     ['Gas', 100],
        #     ['Food', 300],
        #     ['Gym', 50],
        # )
        expenses = (list_3())

        # Start from the first cell. Rows and columns are zero indexed.
        # row = 1
        # col = 0

        row_1 = 1

        # Iterate over the data and write it out row by row.
        # for item, cost in (expenses):
        #     worksheet.write(row, col, item)
        #     worksheet.write(row, col + 1, cost)
        #     row += 1

        for item in (list_3()):
            col_1 = 1
            for item_1 in item:
                worksheet.write(row_1, col_1, item_1)
                print(row_1, col_1, list_2())
                # if (row > col):
                col_1 += 1
            # worksheet.write(row, col, item)
            row_1 += 1

        # # Write a total using a formula.
        # worksheet.write(row, 0, 'Total')
        # worksheet.write(row, 1, '=SUM(B1:B4)')

        workbook.close()
